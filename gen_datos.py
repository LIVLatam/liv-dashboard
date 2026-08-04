#!/usr/bin/env python3
"""Regenera datos.js desde la pestaña 'Datos_Dashboard' del Excel.
Fuente de verdad = Excel. Uso: python3 gen_datos.py
Requiere: openpyxl  (pip install openpyxl --break-system-packages)"""
import json, sys, os
import openpyxl

XLSX = os.environ.get("LIV_XLSX",
    "/sessions/keen-compassionate-carson/mnt/Mi unidad/Logística LIV México - campaña 1.xlsx")
OUT = os.environ.get("LIV_DATOS",
    "/sessions/keen-compassionate-carson/mnt/Mi unidad/Claude · LIV Latam/liv-dashboard-web/datos.js")

def yn(v): return 1 if v == 'Sí' else 0
def n(v):  return 0 if v in (None, '') else (int(v) if float(v) == int(v) else float(v))

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Datos_Dashboard"]
    hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    C, BON = [], {}
    for r in range(2, ws.max_row + 1):
        row = dict(zip(hdr, [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]))
        if row.get('ID') in (None, ''): continue
        cid = int(row['ID'])
        cob = 'Familia' if str(row.get('Cobro') or '').startswith('Famili') else 'Colegio'
        c = {'id': cid, 'n': row['Colegio'], 'es': row['Estado'], 'ci': row['Ciudad'],
             'mu': row.get('Municipio') or '', 'co': row.get('Colonia') or '',
             'pe1': n(row['PE 3']), 'pe2': n(row['PE 4']), 'pe3': n(row['PE 5']),
             'pr1': n(row['PR 1º']), 'pr2': n(row['PR 2º']), 'pr3': n(row['PR 3º']),
             'pr4': n(row['PR 4º']), 'pr5': n(row['PR 5º']), 'pr6': n(row['PR 6º']),
             'pf': n(row['Pre-First']), 'p': n(row['Precio año 1']),
             'bon': row.get('Bonificación (detalle)') or '', 'm': n(row['Facturación vigente']),
             'vig': row.get('Vigencia') or '', 'pago': row.get('Modalidad pago') or '',
             'cob': cob, 'dir': row.get('Dirección') or '',
             'aula': yn(row.get('Aula LIV')), 'f': yn(row.get('Firmado'))}
        # opcionales
        mc = n(row['Facturación contrato'])
        if mc != c['m']: c['mc'] = mc
        ope, opr = row.get('OC alumnos preescolar'), row.get('OC alumnos primaria')
        if ope not in (None, '') or opr not in (None, ''):
            c['occ'] = yn(row.get('OC completa'))
            if ope not in (None, ''): c['ocpe'] = n(ope)
            if opr not in (None, ''): c['ocpr'] = n(opr)
        # ordenar claves como el original
        order = ['id','n','es','ci','mu','co','pe1','pe2','pe3','pr1','pr2','pr3','pr4','pr5','pr6',
                 'pf','p','bon','m','mc','vig','pago','cob','dir','occ','ocpe','ocpr','aula','f']
        C.append({k: c[k] for k in order if k in c})
        BON[str(cid)] = {'y': n(row['Duración (años)']), 'a1': n(row['Precio año 1']),
                         'a2': row.get('Precio año 2') or '—', 'a3': row.get('Precio año 3') or '—',
                         'a4': row.get('Precio año 4') or '—', 'b': row.get('Bonificación año 1') or '',
                         'c': 'Familias' if cob == 'Familia' else 'Colegio',
                         'p': row.get('Particularidad') or ''}
    js = ("/* LIV · Datos del dashboard — generado por gen_datos.py desde la pestaña "
          "'Datos_Dashboard' del Excel. NO editar a mano. */\n"
          "window.LIV_C=" + json.dumps(C, ensure_ascii=False) + ";\n"
          "window.LIV_BON=" + json.dumps(BON, ensure_ascii=False) + ";\n")
    open(OUT, 'w', encoding='utf-8').write(js)
    print(f"OK · {len(C)} colegios → {OUT}")

if __name__ == "__main__":
    main()
